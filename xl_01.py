import openpyxl

def main():
    MyExcel = openpyxl.load_workbook(filename="C:\\Users\\kitri\\Desktop\\x1.xlsx")
    print(MyExcel)
    # for i in MyExcel:
    #     print (i)

    print("\n--------------------------\n")

     #sheet 목록을 확인하는 함수
    sheetName = MyExcel.sheetnames      #data type list
    print(sheetName)

    print("\n--------------------------\n")

    #cell에 접근 1
    v = MyExcel[sheetName[0]]   #sheet1
    print(v['A1'].value)        #A1

    print("\n--------------------------\n")

    #cell에 접근 2
    v2 = v.cell(row=1, column=1).value
    print(v2)

    print("\n--------------------------\n")

    #반복문을 통해 접근하기
    for i in range (1,9):
        val = v["A"+ str(i)].value
        if val is not None:
            print(val, end=" ")

    print("\n--------------------------\n")

    #범위에 접근하기
    for row in v['A1':'B7']:
        s = 0
        for col in row:
            s += col.value
            print(col.value, end=" ")

        print(" = " + str(s))

    print("\n--------------------------\n")

    #해당하는 sheet 전체 데이터 값에 접근하는 방법
    r = v.rows
    for i in r:
        for j in i:
            print(j.value, end="\t")
        print()

if __name__ == "__main__":
    main()